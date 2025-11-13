#!/usr/bin/env python3
import math
import time
import argparse
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date, timedelta

import requests
from azure.identity import DefaultAzureCredential
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

API_MGMT = "https://management.azure.com"
API_VERSION_MG = "2020-05-01"          # management groups descendants
API_VERSION_SUBS_RG = "2021-04-01"     # list resource groups
API_VERSION_BUDGETS = "2023-05-01"     # Microsoft.Consumption/budgets
API_VERSION_COST = "2023-03-01"        # Cost Management query

BUDGET_HEADROOM_PCT = 0.10
BUDGET_ROUND_TO = 100


def log(msg: str, verbose: bool = True) -> None:
    """Simple timestamped log helper."""
    if verbose:
        print(f"[{datetime.now().isoformat()}] {msg}")


def budget_accuracy(budget: float, actual: float) -> float:
    if budget <= 0 and actual <= 0:
        return 1.0
    if budget <= 0 or actual <= 0:
        return 0.0
    return 1.0 - abs(budget - actual) / max(budget, actual)


def get_token(credential: DefaultAzureCredential) -> str:
    return credential.get_token("https://management.azure.com/.default").token


def backoff_sleep(retry_after_header: Optional[str], attempt: int) -> None:
    if retry_after_header:
        try:
            secs = int(retry_after_header)
            time.sleep(max(1, min(secs, 60)))
            return
        except ValueError:
            pass
    time.sleep(min(2 ** attempt, 30))


def do_request(method: str, url: str, token: str, **kwargs) -> requests.Response:
    headers = kwargs.pop("headers", {})
    headers["Authorization"] = f"Bearer {token}"
    headers.setdefault("Content-Type", "application/json")
    attempt = 0
    while True:
        try:
            resp = requests.request(method, url, headers=headers, timeout=60, **kwargs)
        except requests.exceptions.ReadTimeout:
            attempt += 1
            if attempt > 6:
                raise
            backoff_sleep(None, attempt)
            continue

        if resp.status_code in (429, 500, 502, 503, 504):
            attempt += 1
            if attempt > 6:
                resp.raise_for_status()
            backoff_sleep(resp.headers.get("Retry-After"), attempt)
            continue
        resp.raise_for_status()
        return resp


# --------------------- discovery ---------------------
def list_mgs_and_subs_under_mg(mg_id: str, token: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Returns:
      - list of management groups (descendants of mg_id)
      - list of subscriptions (descendants of mg_id)
    """
    url = f"{API_MGMT}/providers/Microsoft.Management/managementGroups/{mg_id}/descendants?api-version={API_VERSION_MG}"
    mgs: Dict[str, Dict[str, Any]] = {}
    subs: Dict[str, Dict[str, Any]] = {}

    while True:
        resp = do_request("GET", url, token)
        payload = resp.json()
        for item in payload.get("value", []):
            t = (item.get("type") or "").lower()
            name = item.get("name")
            props = item.get("properties", {}) or {}
            display_name = props.get("displayName", "")
            if t.endswith("/subscriptions"):
                if name:
                    subs[name] = {
                        "subscriptionId": name,
                        "displayName": display_name,
                        "tenantId": props.get("tenantId", ""),
                    }
            elif t.endswith("/managementgroups"):
                if name:
                    mgs[name] = {
                        "managementGroupId": name,
                        "displayName": display_name,
                    }
        url = payload.get("nextLink") or None
        if not url:
            break

    # Ensure the root MG is included even if descendants API did not return it
    if mg_id not in mgs:
        mgs[mg_id] = {
            "managementGroupId": mg_id,
            "displayName": mg_id,
        }

    return list(mgs.values()), list(subs.values())


def list_resource_groups(sub_id: str, token: str) -> List[str]:
    url = f"{API_MGMT}/subscriptions/{sub_id}/resourcegroups?api-version={API_VERSION_SUBS_RG}"
    rgs: List[str] = []
    while True:
        resp = do_request("GET", url, token)
        j = resp.json()
        for rg in j.get("value", []):
            name = rg.get("name")
            if name:
                rgs.append(name)
        url = j.get("nextLink") or None
        if not url:
            break
    return rgs


# --------------------- budgets ---------------------
def list_budgets_at_scope(scope: str, token: str) -> List[Dict[str, Any]]:
    url = f"{API_MGMT}{scope}/providers/Microsoft.Consumption/budgets?api-version={API_VERSION_BUDGETS}"
    resp = do_request("GET", url, token)
    return resp.json().get("value", []) or []


# --------------------- cost and forecast ---------------------
def first_last_day_of_last_n_months(n: int = 3) -> List[Tuple[date, date]]:
    today = date.today()
    first_this_month = today.replace(day=1)
    month_cursor = first_this_month - relativedelta(months=1)
    out: List[Tuple[date, date]] = []
    for _ in range(n):
        start = month_cursor.replace(day=1)
        end = (start + relativedelta(months=1)) - timedelta(days=1)
        out.append((start, end))
        month_cursor = start - relativedelta(months=1)
    return out


def cost_query_last_months_scope(scope: str, months: int, token: str) -> List[float]:
    periods = first_last_day_of_last_n_months(months)
    costs: List[float] = []
    for (start, end) in periods:
        url = f"{API_MGMT}{scope}/providers/Microsoft.CostManagement/query?api-version={API_VERSION_COST}"
        body = {
            "type": "Usage",
            "dataSet": {
                "granularity": "None",
                "aggregation": {"totalCost": {"name": "Cost", "function": "Sum"}},
            },
            "timeframe": "Custom",
            "timePeriod": {
                "from": f"{start.isoformat()}T00:00:00Z",
                "to": f"{(end + timedelta(days=1)).isoformat()}T00:00:00Z",
            },
        }
        resp = do_request("POST", url, token, json=body)
        j = resp.json()
        val = 0.0
        for row in j.get("properties", {}).get("rows", []):
            try:
                val += float(row[0] or 0.0)
            except (ValueError, TypeError):
                pass
        costs.append(round(val, 2))
    return costs


def cost_forecast_current_month_scope(scope: str, token: str) -> Optional[float]:
    url = f"{API_MGMT}{scope}/providers/Microsoft.CostManagement/query?api-version={API_VERSION_COST}"
    body = {
        "type": "Usage",
        "timeframe": "MonthToDate",
        "dataSet": {
            "granularity": "Daily",
            "aggregation": {"totalCost": {"name": "Cost", "function": "Sum"}},
        },
        "includeForecast": True,
    }
    resp = do_request("POST", url, token, json=body)
    j = resp.json()
    cols = [c["name"] for c in j.get("properties", {}).get("columns", [])]
    rows = j.get("properties", {}).get("rows", [])

    try:
        idx_cost = cols.index("Cost")
        idx_is_fc = cols.index("IsForecast")
    except ValueError:
        total = 0.0
        for r in rows:
            try:
                total += float(r[0] or 0.0)
            except (ValueError, TypeError):
                pass
        return round(total, 2)

    total = 0.0
    for r in rows:
        try:
            if str(r[idx_is_fc]).lower() in ("true", "1", "yes"):
                total += float(r[idx_cost] or 0.0)
        except (ValueError, TypeError):
            pass
    return round(total, 2)


# --------------------- misc ---------------------
def compute_suggested_budget(value: float, prev_two_months: List[float]) -> float:
    baseline = max(value, (value + sum(prev_two_months)) / (1 + len(prev_two_months)))
    suggestion = baseline * (1.0 + BUDGET_HEADROOM_PCT)
    if BUDGET_ROUND_TO > 0:
        suggestion = math.ceil(suggestion / BUDGET_ROUND_TO) * BUDGET_ROUND_TO
    return round(suggestion, 2)


def flatten_notifications(notifs: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if not notifs:
        return out
    for key, cfg in notifs.items():
        if not isinstance(cfg, dict):
            continue
        out.append(
            {
                "ConditionKey": key,
                "Enabled": bool(cfg.get("enabled", True)),
                "ThresholdType": cfg.get("thresholdType", ""),
                "Operator": cfg.get("operator", ""),
                "ThresholdPercent": cfg.get("threshold", None),
                "ContactEmails": ";".join(cfg.get("contactEmails", []) or []),
                "ContactGroups": ";".join(cfg.get("contactGroups", []) or []),
                "ContactRoles": ";".join(cfg.get("contactRoles", []) or []),
            }
        )
    out.sort(
        key=lambda r: (
            r.get("ThresholdType", ""),
            float(r.get("ThresholdPercent") or 0),
        )
    )
    return out


def parse_args():
    p = argparse.ArgumentParser(
        description="Assess Azure budgets at MG, Sub, and RG scopes and suggest improvements."
    )
    p.add_argument("management_group_id", help="Root Management Group ID (for example mg-demo)")
    p.add_argument("--out", default=None, help="Output Excel path (.xlsx)")
    p.add_argument(
        "--months",
        type=int,
        default=3,
        help="Number of full past months to compare (excludes current month)",
    )
    p.add_argument(
        "--scopes",
        default="mg,sub,rg",
        help="Comma list of scopes to include: mg,sub,rg",
    )
    p.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose progress logging",
    )
    p.add_argument(
        "--subscription-ids",
        nargs="+",
        help=(
            "Optional list of subscription IDs to include. "
            "If specified, only these subscriptions will be processed for sub/rg scopes."
        ),
    )
    p.add_argument(
        "--rg-names",
        nargs="+",
        help=(
            "Optional list of resource group names to include. "
            "If specified, only RGs with these names will be processed in each included subscription."
        ),
    )
    return p.parse_args()


def main():
    args = parse_args()
    root_mg_id = args.management_group_id
    months = args.months
    include_scopes = {s.strip().lower() for s in args.scopes.split(",") if s.strip()}
    out_path = args.out or f"budget_assessment_{root_mg_id}_{date.today().isoformat()}.xlsx"
    verbose = args.verbose

    # Normalise filters
    sub_filter = {s.lower() for s in (args.subscription_ids or [])}
    rg_filter = {r.lower() for r in (args.rg_names or [])}

    cred = DefaultAzureCredential(exclude_interactive_browser_credential=False)
    token = get_token(cred)

    log("Enumerating management groups and subscriptions...", verbose)
    mgs, subs = list_mgs_and_subs_under_mg(root_mg_id, token) if (
        "mg" in include_scopes or "sub" in include_scopes or "rg" in include_scopes
    ) else ([], [])

    # Apply subscription filter if provided
    if sub_filter:
        original_count = len(subs)
        subs = [s for s in subs if s["subscriptionId"].lower() in sub_filter]
        missing = sub_filter - {s["subscriptionId"].lower() for s in subs}
        print(f"  Filtered subscriptions by --subscription-ids: {len(subs)} of {original_count} remain.")
        if missing:
            print("  Warning: the following subscription IDs were not found under this management group:")
            for mid in sorted(missing):
                print(f"    - {mid}")

    print(f"  Management Groups (including root and descendants): {len(mgs)}")
    print(f"  Subscriptions: {len(subs)}")

    mg_rows: List[Dict[str, Any]] = []
    sub_rows: List[Dict[str, Any]] = []
    sub_nobudget_rows: List[Dict[str, Any]] = []
    rg_rows: List[Dict[str, Any]] = []

    # Management group budgets
    if "mg" in include_scopes:
        log("Processing management group budgets...", verbose)
        for idx, mg in enumerate(mgs, 1):
            mg_id = mg["managementGroupId"]
            mg_scope = f"/providers/Microsoft.Management/managementGroups/{mg_id}"
            log(f"MG {idx}/{len(mgs)} - {mg_id}", verbose)

            try:
                mg_budgets = list_budgets_at_scope(mg_scope, token)
            except Exception as e:
                print(f"  ! MG budgets fetch failed ({mg_id}): {e}")
                mg_budgets = []

            if not mg_budgets:
                continue

            try:
                mg_monthly = cost_query_last_months_scope(mg_scope, months, token)
            except Exception as e:
                print(f"  ! MG cost query failed ({mg_id}): {e}")
                mg_monthly = [None] * months

            mg_last = mg_monthly[0] if mg_monthly and mg_monthly[0] is not None else None
            mg_prev2 = [c for c in mg_monthly[1:3] if c is not None]

            try:
                mg_fc = cost_forecast_current_month_scope(mg_scope, token)
            except Exception as e:
                print(f"  ! MG forecast failed ({mg_id}): {e}")
                mg_fc = None

            for b in mg_budgets:
                emit_budget_rows(
                    rows_list=mg_rows,
                    scope_type="MG",
                    scope_id=mg_scope,
                    sub_name="",
                    sub_id="",
                    budget=b,
                    last_val=mg_last,
                    prev_two=mg_prev2,
                    forecast_total=mg_fc,
                )

    # Subscription and RG budgets
    if "sub" in include_scopes or "rg" in include_scopes:
        log("Processing subscriptions...", verbose)

    for s_idx, s in enumerate(subs, 1):
        sub_id = s["subscriptionId"]
        sub_name = s.get("displayName", "")
        sub_scope = f"/subscriptions/{sub_id}"

        log(f"Subscription {s_idx}/{len(subs)} - {sub_name} ({sub_id})", verbose)

        # Subscription budgets and missing budget dedemotion
        if "sub" in include_scopes:
            try:
                sub_budgets = list_budgets_at_scope(sub_scope, token)
            except Exception as e:
                print(f"  ! Sub budgets fetch failed ({sub_id}): {e}")
                sub_budgets = []

            try:
                sub_monthly = cost_query_last_months_scope(sub_scope, months, token)
            except Exception as e:
                print(f"  ! Sub cost query failed ({sub_id}): {e}")
                sub_monthly = [None] * months

            sub_last = sub_monthly[0] if sub_monthly and sub_monthly[0] is not None else None
            sub_prev2 = [c for c in sub_monthly[1:3] if c is not None]

            try:
                sub_fc = cost_forecast_current_month_scope(sub_scope, token)
            except Exception as e:
                print(f"  ! Sub forecast failed ({sub_id}): {e}")
                sub_fc = None

            if sub_budgets:
                for b in sub_budgets:
                    emit_budget_rows(
                        rows_list=sub_rows,
                        scope_type="Subscription",
                        scope_id=sub_scope,
                        sub_name=sub_name,
                        sub_id=sub_id,
                        budget=b,
                        last_val=sub_last,
                        prev_two=sub_prev2,
                        forecast_total=sub_fc,
                    )
            else:
                suggested_actual_based = (
                    compute_suggested_budget(sub_last, sub_prev2)
                    if sub_last is not None
                    else ""
                )
                suggested_forecast_based = (
                    compute_suggested_budget(sub_fc, sub_prev2)
                    if sub_fc is not None
                    else ""
                )

                sub_nobudget_rows.append(
                    {
                        "ScopeType": "Subscription",
                        "ScopeId": sub_scope,
                        "SubscriptionName": sub_name,
                        "SubscriptionId": sub_id,
                        "ResourceGroup": "",
                        "BudgetName": "",
                        "BudgetAmount": "",
                        "BudgetTimeGrain": "",
                        "BudgetStartDate": "",
                        "BudgetEndDate": "",
                        "ConditionKey": "",
                        "ThresholdType": "",
                        "Operator": "",
                        "ThresholdPercent": "",
                        "Enabled": "",
                        "ContactEmails": "",
                        "ContactGroups": "",
                        "ContactRoles": "",
                        "LastMonthCost": sub_last if sub_last is not None else "",
                        "PrevMonthCost": sub_prev2[0] if len(sub_prev2) > 0 else "",
                        "Prev2MonthCost": sub_prev2[1] if len(sub_prev2) > 1 else "",
                        "PercentOfBudgetLastMonth": "",
                        "BudgetAccuracy": "",
                        "CurrentMonthForecastTotal": sub_fc if sub_fc is not None else "",
                        "ForecastPercentOfBudget": "",
                        "ForecastConditionWillTrigger": "",
                        "SuggestedBudget_ActualBased": suggested_actual_based,
                        "SuggestedBudget_ForecastBased": suggested_forecast_based,
                        "SuggestionNote": "No budget configured at this scope. Suggestions = max(value, 3-mo avg) + 10 percent headroom, rounded.",
                    }
                )

        # Resource group budgets
        if "rg" in include_scopes:
            try:
                rgs = list_resource_groups(sub_id, token)
                if rg_filter:
                    orig_rg_count = len(rgs)
                    rgs = [rg for rg in rgs if rg.lower() in rg_filter]
                    log(
                        f"  Found {orig_rg_count} resource groups in subscription {sub_name}, "
                        f"{len(rgs)} match --rg-names filter",
                        verbose,
                    )
                else:
                    log(f"  Found {len(rgs)} resource groups in subscription {sub_name}", verbose)
            except Exception as e:
                print(f"  ! RG list failed ({sub_id}): {e}")
                rgs = []

            for rg in rgs:
                rg_scope = f"/subscriptions/{sub_id}/resourceGroups/{rg}"
                try:
                    rg_budgets = list_budgets_at_scope(rg_scope, token)
                except Exception as e:
                    print(f"  ! RG budgets fetch failed ({sub_id}/{rg}): {e}")
                    rg_budgets = []

                if not rg_budgets:
                    continue

                try:
                    rg_monthly = cost_query_last_months_scope(rg_scope, months, token)
                except Exception as e:
                    print(f"  ! RG cost query failed ({sub_id}/{rg}): {e}")
                    rg_monthly = [None] * months

                rg_last = rg_monthly[0] if rg_monthly and rg_monthly[0] is not None else None
                rg_prev2 = [c for c in rg_monthly[1:3] if c is not None]

                try:
                    rg_fc = cost_forecast_current_month_scope(rg_scope, token)
                except Exception as e:
                    print(f"  ! RG forecast failed ({sub_id}/{rg}): {e}")
                    rg_fc = None

                for b in rg_budgets:
                    emit_budget_rows(
                        rows_list=rg_rows,
                        scope_type="ResourceGroup",
                        scope_id=rg_scope,
                        sub_name=sub_name,
                        sub_id=sub_id,
                        budget=b,
                        last_val=rg_last,
                        prev_two=rg_prev2,
                        forecast_total=rg_fc,
                        resource_group=rg,
                    )

    # write Excel with separate tabs
    fieldnames = [
        "ScopeType",
        "ScopeId",
        "SubscriptionName",
        "SubscriptionId",
        "ResourceGroup",
        "BudgetName",
        "BudgetAmount",
        "BudgetTimeGrain",
        "BudgetStartDate",
        "BudgetEndDate",
        "ConditionKey",
        "ThresholdType",
        "Operator",
        "ThresholdPercent",
        "Enabled",
        "ContactEmails",
        "ContactGroups",
        "ContactRoles",
        "LastMonthCost",
        "PrevMonthCost",
        "Prev2MonthCost",
        "PercentOfBudgetLastMonth",
        "BudgetAccuracy",
        "CurrentMonthForecastTotal",
        "ForecastPercentOfBudget",
        "ForecastConditionWillTrigger",
        "SuggestedBudget_ActualBased",
        "SuggestedBudget_ForecastBased",
        "SuggestionNote",
    ]

    wb = Workbook()
    created_any = False

    def create_sheet(title: str, rows: List[Dict[str, Any]]) -> None:
        nonlocal created_any
        if not rows:
            return
        if not created_any:
            ws = wb.active
            ws.title = title
            created_any = True
        else:
            ws = wb.create_sheet(title=title)
        ws.append(fieldnames)
        for r in rows:
            ws.append([r.get(col, "") for col in fieldnames])

    if "mg" in include_scopes:
        create_sheet("MG_Budgets", mg_rows)
    if "sub" in include_scopes:
        create_sheet("Sub_Budgets", sub_rows)
        create_sheet("Sub_NoBudget", sub_nobudget_rows)
    if "rg" in include_scopes:
        create_sheet("RG_Budgets", rg_rows)

    wb.save(out_path)
    print(f"\nDone. Wrote Excel workbook: {out_path}")


def emit_budget_rows(
    rows_list: List[Dict[str, Any]],
    scope_type: str,
    scope_id: str,
    sub_name: str,
    sub_id: str,
    budget: Dict[str, Any],
    last_val: Optional[float],
    prev_two: List[float],
    forecast_total: Optional[float],
    resource_group: str = "",
):
    props = budget.get("properties", {}) or {}
    budget_name = budget.get("name", "")
    amount = props.get("amount", None)
    time_grain = props.get("timeGrain", "")
    tp = props.get("timePeriod", {}) or {}
    start_date = tp.get("startDate", "")
    end_date = tp.get("endDate", "")

    conditions = flatten_notifications(props.get("notifications")) or [None]

    acc = (
        round(budget_accuracy(float(amount), float(last_val)), 4)
        if (last_val is not None and amount)
        else ""
    )

    suggested_actual_based = (
        compute_suggested_budget(last_val, prev_two) if last_val is not None else ""
    )
    suggested_forecast_based = (
        compute_suggested_budget(forecast_total, prev_two)
        if forecast_total is not None
        else ""
    )

    percent_of_budget_last_month = (
        round((last_val / float(amount)) * 100, 2)
        if (last_val and amount)
        else ""
    )
    forecast_percent_of_budget = (
        round((forecast_total / float(amount)) * 100, 2)
        if (forecast_total and amount)
        else ""
    )

    for cond in conditions:
        will_trigger = ""
        if cond and cond.get("ThresholdType", "").lower().startswith("forecast"):
            if forecast_percent_of_budget != "" and cond.get("ThresholdPercent") is not None:
                try:
                    will_trigger = str(
                        float(forecast_percent_of_budget)
                        >= float(cond["ThresholdPercent"])
                    )
                except Exception:
                    will_trigger = ""

        rows_list.append(
            {
                "ScopeType": scope_type,
                "ScopeId": scope_id,
                "SubscriptionName": sub_name,
                "SubscriptionId": sub_id,
                "ResourceGroup": resource_group,
                "BudgetName": budget_name,
                "BudgetAmount": round(float(amount), 2) if amount is not None else "",
                "BudgetTimeGrain": time_grain,
                "BudgetStartDate": start_date,
                "BudgetEndDate": end_date or "",
                "ConditionKey": cond.get("ConditionKey", "") if cond else "",
                "ThresholdType": cond.get("ThresholdType", "") if cond else "",
                "Operator": cond.get("Operator", "") if cond else "",
                "ThresholdPercent": cond.get("ThresholdPercent", "") if cond else "",
                "Enabled": cond.get("Enabled", "") if cond else "",
                "ContactEmails": cond.get("ContactEmails", "") if cond else "",
                "ContactGroups": cond.get("ContactGroups", "") if cond else "",
                "ContactRoles": cond.get("ContactRoles", "") if cond else "",
                "LastMonthCost": last_val if last_val is not None else "",
                "PrevMonthCost": prev_two[0] if len(prev_two) > 0 else "",
                "Prev2MonthCost": prev_two[1] if len(prev_two) > 1 else "",
                "PercentOfBudgetLastMonth": percent_of_budget_last_month,
                "BudgetAccuracy": acc,
                "CurrentMonthForecastTotal": forecast_total if forecast_total is not None else "",
                "ForecastPercentOfBudget": forecast_percent_of_budget,
                "ForecastConditionWillTrigger": will_trigger,
                "SuggestedBudget_ActualBased": suggested_actual_based,
                "SuggestedBudget_ForecastBased": suggested_forecast_based,
                "SuggestionNote": "Suggestions = max(value, 3-mo avg) + 10 percent headroom, rounded.",
            }
        )


if __name__ == "__main__":
    main()
